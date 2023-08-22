from openpyxl import Workbook, load_workbook

def criar_planilha():
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Comandas"
    planilha.append(["Número da Comanda", "Produtos", "Quantidade", "Preço"])

    return workbook

def nova_comanda(planilha):
    num_mesa =input("Digite o número da nova comanda: ")
    num_comanda = "Mesa " + num_mesa
    planilha.create_sheet(title=num_comanda)
    print(f"{num_comanda} criada com sucesso.")

def adicionar_produtos(planilha):
    num_comanda = input("Digite o número da comanda: ")
    if num_comanda in planilha.sheetnames:
        comanda = planilha[num_comanda]
        linha_atual = comanda.max_row
        if linha_atual == 1:
            comanda.cell(row=linha_atual, column=1, value="Produtos")
            comanda.cell(row=linha_atual, column=2, value="Preço")
            comanda.cell(row=linha_atual, column=3, value="Quant")
            comanda.cell(row=linha_atual, column=4, value="Valor R$")
            linha_atual += 2
        
        while True:
            print("-----------------------------")
            print("Deseja adicionar algum pedido?")
            print("[1] - Sim.")
            print("[2] - Não.")
            print("-----------")
            escolha = int(input("Escolha uma opção: "))
            
            if escolha == 1:
                Produto = input("Digite o produto a ser adicionado: ")
                Preco = float(input("Digite o preço: "))
                Quantidade = int(input("Digite a quantidade: "))
                Valor = Quantidade * Preco
                
                comanda.cell(row=linha_atual, column=1, value=Produto)
                comanda.cell(row=linha_atual, column=2, value=Preco)
                comanda.cell(row=linha_atual, column=3, value=Quantidade)
                comanda.cell(row=linha_atual, column=4, value=Valor)
                
                linha_atual += 1
            else:
                print("Produtos salvos com sucesso")
                break
    else:
        print("Comanda não encontrada.")

def fechar_comanda(planilha):
    num_comanda = input("Digite o número da comanda a ser fechada: ")
    if num_comanda in planilha.sheetnames:
        produtos = planilha[num_comanda].cell(row=2, column=1).value
        print(f"{num_comanda} contém os seguintes produtos: {produtos}")
        planilha.remove(planilha[num_comanda])
        print("Comanda fechada.")
    else:
        print("Comanda não encontrada.")

def salvar_planilha(workbook):
    nome_arquivo = "comandas.xlsx"
    workbook.save(nome_arquivo)
    print(f"Planilha salva como '{nome_arquivo}'.")

def menu():
    workbook = criar_planilha()

    while True:
        print("\n--- Menu ---")
        print("[1] - Nova Comanda")
        print("[2] - Adicionar Produtos")
        print("[3] - Fechar Comanda")
        print("[4] - Salvar Planilha")
        print("[5] - Sair do Programa")

        opcao = input("Escolha uma opção: ")

        if opcao == "1":
            nova_comanda(workbook)
        elif opcao == "2":
            adicionar_produtos(workbook)
        elif opcao == "3":
            fechar_comanda(workbook)
        elif opcao == "4":
            salvar_planilha(workbook)
        elif opcao == "5":
            print("Saindo do programa...")
            break
        else:
            print("Opção inválida. Por favor, escolha uma opção válida.")

if __name__ == "__main__":
    menu()